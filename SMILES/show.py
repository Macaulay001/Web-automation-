import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Select the sheet you want to work with
worksheet = workbook['Sheet1']

# Keep track of the non-empty rows in column F
non_empty_rows = []

# Loop through all rows in the worksheet
for row in worksheet.iter_rows(min_row=1, values_only=True):
    # If the cell in column F is empty, skip the row
    if not row[5]:
        continue
    # Otherwise, add the row to the non-empty rows list
    non_empty_rows.append(row)

# Get the value of the first row in the worksheet to use as the cleaned data sheet name
cleaned_data_sheet_name = worksheet[1][0].value

# Create a new worksheet to store the cleaned up version of the data
cleaned_worksheet = workbook.create_sheet(cleaned_data_sheet_name)

# Copy the entire row for each non-empty row in column F to the new worksheet
for row in non_empty_rows:
    cleaned_worksheet.append(row)

# Save the cleaned up version of the workbook to the new sheet
cleaned_worksheet.title = cleaned_data_sheet_name
workbook.save('example_cleaned.xlsx')


for row in non_empty_rows:
    cleaned_row = []
    for cell in row:
        if cell.data_type == 's':
            cleaned_row.append(cell.value)
        else:
            cleaned_row.append(cell.value)
            # Update the formula bar links to match the rows and columns format of the new sheet
            if cell.data_type == 'f':
                cleaned_cell = cleaned_worksheet.cell(row=len(non_empty_rows)+1, column=cell.column)
                cleaned_cell.value = cell.value
                cleaned_cell.coordinate = cleaned_cell.column_letter + str(len(non_empty_rows)+1)
    cleaned_worksheet.append(cleaned_row)

for cell in row:
    if isinstance(cell, openpyxl.cell.cell.Cell) and cell.data_type == 'f':
        cleaned_cell = cleaned_worksheet.cell(row=len(non_empty_rows)+1, column=cell.column)
        cleaned_cell.value = cell.value
        cleaned_cell.coordinate = cleaned_cell.column_letter + str(len(non_empty_rows)+1)
    cleaned_row.append(cell.value)
